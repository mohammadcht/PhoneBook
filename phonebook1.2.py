from openpyxl import Workbook, load_workbook

app_name, app_ver = 'PhoneNumber', '1.2'
first_names, last_names, phone_numbers, index_list = [], [], [], []
count, found = 1, None

print(f'\n>>> Welcome to {app_name} V{app_ver} <<<')
while True :
    print('''\n++++++++++++++++++++++++
** Enter 0 to Save & exit
** Enter 1 to load contacts from Excel

=> Enter 2 to add a Contact
=> Enter 3 to search in Contacts
=> Enter 4 to Display all Contacts
=> Enter 5 to Delete a Contact
++++++++++++++++++++++++''')
    print(f'Number of Contacts : {count-1}\n++++++++++++++++++++++++')
    menu = (input('> '))
    if menu.isalpha():
        print('digit only, letters are not allowed :( ')
        continue
    else:
        menu=int(menu)
    if menu == 0 :
        count1 = count
        while count1 > 0:
            sheet.delete_cols(idx=count1), sheet.delete_rows(idx=count1)
            count1 -= 1 
        save_count_f, save_count_l, save_count_p = 1, 1, 1
        for i in first_names :
            sheet_a = 'A'+str(save_count_f)
            sheet[sheet_a] = i
            workbook.save(filename='phonebook.xlsx')
            save_count_f += 1
        for i in last_names :
            sheet_b = 'B'+str(save_count_l)
            sheet[sheet_b] = i
            workbook.save(filename='phonebook.xlsx')
            save_count_l += 1
        for i in phone_numbers :
            sheet_c = 'C'+str(save_count_p)
            sheet[sheet_c] = i
            workbook.save(filename='phonebook.xlsx')
            save_count_p += 1
        break
    elif menu == 1 :
        try :
            workbook = load_workbook(filename='phonebook.xlsx')
            sheet = workbook["Sheet"]
            for row in sheet.iter_rows(values_only=True):
                first_names.append(row[0]), last_names.append(row[1]), phone_numbers.append(row[2])
                count += 1
            if count == 2 :
                print('1 Contact aded!')
            elif count > 2 :
                print(f'{count-1} Contacts aded!')
            workbook.close()
        except :
            print('''
___________________________
| phonebook.xlsx not found |
###########################
''')
    elif menu == 2 :
        workbook = Workbook()
        sheet = workbook.active
        first_name, last_name, phone_number = input('first Name :\n>'), input('last Name :\n>'), input('Phone Number :\n>')
        first_names.append(first_name), last_names.append(last_name), phone_numbers.append(phone_number)
        count += 1
    elif menu == 3 :
        search_term = input('\nEnter name or phone to search :\n>')
        s_first_names, s_last_names, s_phone_numbers = [], [], []
        count_n = 1
        s_first_names,s_last_names, s_phone_numbers = first_names.copy(), last_names.copy(), phone_numbers.copy()
        print('Search Result:')
        for i in s_first_names :
            if search_term in s_first_names:
                index = s_first_names.index(search_term)
                last_name, phone_number = s_last_names[index], s_phone_numbers[index]
                print(f'{count_n}- {search_term} {last_name} {phone_number}')
                s_first_names.remove(s_first_names[index]), s_last_names.remove(s_last_names[index]), s_phone_numbers.remove(s_phone_numbers[index])
                count_n += 1
        for i in s_last_names :
            if search_term in s_last_names:
                index = s_last_names.index(search_term)
                first_name, phone_number = s_first_names[index], s_phone_numbers[index]
                print(f'{count_n}- {first_name} {search_term} {phone_number}')
                s_first_names.remove(s_first_names[index]), s_last_names.remove(s_last_names[index]), s_phone_numbers.remove(s_phone_numbers[index])
                count_n += 1
        for i in s_phone_numbers :
            if search_term in s_phone_numbers:
                index = s_phone_numbers.index(search_term)
                last_name, first_name = s_last_names[index], s_first_names[index]
                print(f'{count_n}- {first_name} {last_name} {search_term}')
                s_first_names.remove(s_first_names[index]), s_last_names.remove(s_last_names[index]), s_phone_numbers.remove(s_phone_numbers[index])
                count_n += 1
    elif menu == 4 :
        print('\nName\t\t\t\t\tPhone Number\n')
        for i in range(count-1):
            print(f'{i+1}- {first_names[i-1]} {last_names[i-1]}\t\t\t{phone_numbers[i-1]}')
    elif menu == 5 :
        search_term = input('\nEnter name or phone to search :\n>')
        s_first_names, s_last_names, s_phone_numbers, index_list = [], [], [], []
        count_n, count_delete = 1, 0
        s_first_names,s_last_names, s_phone_numbers = first_names.copy(), last_names.copy(), phone_numbers.copy()
        print('Search Result:')
        for i in s_first_names :
            found_f = False
            if search_term in s_first_names:
                index = s_first_names.index(search_term)
                index_count = count_delete + index
                index_list.append(index_count)
                last_name, phone_number = s_last_names[index], s_phone_numbers[index]
                print(f'{count_n}- {search_term} {last_name} {phone_number}')
                s_first_names.remove(s_first_names[index]), s_last_names.remove(s_last_names[index]), s_phone_numbers.remove(s_phone_numbers[index])
                count_n += 1
                count_delete += 1
                found_f = True
        for i in s_last_names :
            found_l = False
            if search_term in s_last_names:
                index = s_last_names.index(search_term)
                index_count = count_delete + index
                index_list.append(index_count)
                first_name, phone_number = s_first_names[index], s_phone_numbers[index]
                print(f'{count_n}- {first_name} {search_term} {phone_number}')
                s_first_names.remove(s_first_names[index]), s_last_names.remove(s_last_names[index]), s_phone_numbers.remove(s_phone_numbers[index])
                count_n += 1
                count_delete += 1
                found_l = True
        for i in s_phone_numbers :
            found_p = False
            if search_term in s_phone_numbers:
                index = s_phone_numbers.index(search_term)
                index_count = count_delete + index
                index_list.append(index_count)
                last_name, first_name = s_last_names[index], s_first_names[index]
                print(f'{count_n}- {first_name} {last_name} {search_term}')
                s_first_names.remove(s_first_names[index]), s_last_names.remove(s_last_names[index]), s_phone_numbers.remove(s_phone_numbers[index])
                count_n += 1
                count_delete += 1
                found_p = True
        while True:    
            if found_f == False and found_l == False and found_p == False:
                print("Not Found")
                break
            else:
                index_num = input('Enter Number of Contact you want delete it: (Enter 0 to Exit)\n>')
                if index_num.isalpha():
                    print('digit only letters are not allowed :( ')
                    continue
                else :
                    index_num = int(index_num)
                    if index_num == 0:
                        break
                    else:
                        index_number = index_num - 1
                        index_id = index_list[index_number]
                        del first_names[index_id],last_names[index_id],phone_numbers[index_id]
                        count -= 1
                        break
    else :
        print('wrong entry')